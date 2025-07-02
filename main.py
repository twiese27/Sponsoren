import pandas as pd
import re
import sqlite3
from openpyxl import load_workbook

def extract_emails(text):
    if pd.isna(text):
        return []
    return re.findall(r'[\w\.-]+@[\w\.-]+', str(text))

def extract_name_and_address_heuristik(text):
    if pd.isna(text) or not str(text).strip():
        return "", ""
    text_ohne_mail = re.sub(r'[\w\.-]+@[\w\.-]+', '', str(text))
    lines = re.split(r'[\n;,]', text_ohne_mail)
    lines = [l.strip() for l in lines if l.strip()]
    if not lines:
        return "", ""
    joined = ' '.join(lines)
    firmenendungen = [
        r'GmbH\s*&\s*Co\.\s*KG', r'GmbH\s*&\s*Co\s*KG', r'GmbH', r'GmBH', r'AG', r'KG', r'e\.K\.', r'e\.K', r'UG', r'OHG', r'e\.V\.', r'e\.V', r'Ltd\.', r'Ltd', r'GbR',
        r'Comp\.', r'Comp', r'&\s*Co\.', r'&\s*Co'
    ]
    last_match = None
    for fe in firmenendungen:
        for m in re.finditer(fe, joined, re.IGNORECASE):
            if (last_match is None) or (m.end() > last_match.end()):
                last_match = m
    if last_match:
        name = joined[:last_match.end()].strip(' ,;\n')
        adresse = joined[last_match.end():].strip(' ,;\n')
        return name, adresse
    street_keywords = [
        'straße', 'str.', 'weg', 'platz', 'allee', 'ring', 'gasse', 'chaussee', 'damm',
        'ufer', 'stieg', 'pfad', 'steig', 'markt', 'tor', 'brücke', 'am ', 'auf der ', 'an der ', 'im ', 'in der ', 'unterer ', 'oberer ', 'hinterm ', 'vor dem', 'zum ', 'zur ', 'bei ', 'bei der '
    ]
    addr_start_idx = None
    for i, line in enumerate(lines):
        has_number = re.search(r'\d', line)
        has_street = any(kw in line.lower() for kw in street_keywords) or re.match(r'^(am|auf der|an der|im|in der|unterer|oberer|hinterm|vor dem|zum|zur|bei|bei der)\b', line.lower())
        if has_number and has_street:
            addr_start_idx = i
            break
    if addr_start_idx is not None:
        name = ' '.join(lines[:addr_start_idx]).strip(' ,;\n')
        adresse = ' '.join(lines[addr_start_idx:]).strip(' ,;\n')
        return name, adresse
    match = re.search(r'\d{1,5}[a-zA-Z]?(\s|$)', joined)
    if match:
        idx = match.start()
        name = joined[:idx].strip(' ,;\n')
        adresse = joined[idx:].strip(' ,;\n')
        return name, adresse
    if len(lines) >= 3:
        name = ' '.join(lines[:-2]).strip(' ,;\n')
        adresse = ' '.join(lines[-2:]).strip(' ,;\n')
        return name, adresse
    elif len(lines) == 2:
        return lines[0], lines[1]
    elif len(lines) == 1:
        return lines[0], ""
    else:
        return "", ""

def bool_and_info(val):
    """Gibt (bool, info) zurück, wenn val ein boolescher Wert ist, sonst Zusatzinfo"""
    if pd.isna(val):
        return 0, ""
    s = str(val).strip().lower()
    # Akzeptiere Varianten für Ja/Nein
    if s in ["ja", "yes", "y", "1", "true"]:
        return 1, ""
    if s in ["nein", "no", "n", "0", "false"]:
        return 0, ""
    # Enthält "ja" oder "nein" am Anfang, dann bool entsprechend, Rest als Info
    if s.startswith("ja"):
        return 1, val
    if s.startswith("nein"):
        return 0, val
    # Sonst: Info, Bool auf 0
    return 0, val

# Excel einlesen (Daten)
df = pd.read_excel("Sponsoring25.xlsx")

# Rot-Markierung mit openpyxl auslesen
wb = load_workbook("Sponsoring25.xlsx")
ws = wb.active
rot_markiert = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    cell = row[0]
    fill = cell.fill
    is_rot = False
    if fill and fill.start_color and fill.start_color.type == "rgb":
        rgb = fill.start_color.rgb
        if rgb and rgb.startswith("FFFF"):
            is_rot = True
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        rgb = fill.fgColor.rgb
        if rgb and rgb.upper().startswith("FFFF"):
            is_rot = True
    rot_markiert.append(is_rot)

# Datenbankverbindung und Tabellen anlegen
conn = sqlite3.connect("sponsoren.db")
cursor = conn.cursor()

cursor.execute("DROP TABLE IF EXISTS sponsor")
cursor.execute("DROP TABLE IF EXISTS mailadresse")

cursor.execute("""
CREATE TABLE IF NOT EXISTS sponsor (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    adresse TEXT,
    wichtig TEXT,
    angefragt INTEGER,
    sponsor_ja_nein INTEGER,
    rueckmeldung INTEGER,
    info TEXT,
    gegenleistung TEXT,
    was TEXT,
    sponsoring_2024 TEXT,
    nie_mehr_anfragen INTEGER
)
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS mailadresse (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sponsor_id INTEGER,
    mail TEXT,
    FOREIGN KEY(sponsor_id) REFERENCES sponsor(id)
)
""")
conn.commit()

for idx, row in df.iterrows():
    sponsor_text = row['Sponsoren']
    emails = extract_emails(sponsor_text)
    name, adresse = extract_name_and_address_heuristik(sponsor_text)
    nie_mehr_anfragen = 1 if idx < len(rot_markiert) and rot_markiert[idx] else 0

    angefragt_bool, angefragt_info = bool_and_info(row.get('Angefragt?', ''))
    sponsor_bool, sponsor_info = bool_and_info(row.get('Sponsor Ja?/Nein?', ''))
    rueckmeldung_bool, rueckmeldung_info = bool_and_info(row.get('Rückmeldung?', ''))

    # Zusatzinfos bündeln, falls vorhanden
    info_parts = []
    if angefragt_info:
        info_parts.append(f"Angefragt: {angefragt_info}")
    if sponsor_info:
        info_parts.append(f"Sponsor Ja/Nein: {sponsor_info}")
    if rueckmeldung_info:
        info_parts.append(f"Rückmeldung: {rueckmeldung_info}")
    info = " | ".join(info_parts)

    cursor.execute("""
        INSERT INTO sponsor (
            name, adresse, wichtig, angefragt, sponsor_ja_nein, rueckmeldung, info, gegenleistung, was, sponsoring_2024, nie_mehr_anfragen
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        name,
        adresse,
        row.get('Wichtig!', ''),
        angefragt_bool,
        sponsor_bool,
        rueckmeldung_bool,
        info,
        row.get('Gegenleistung?', ''),
        row.get('Was?', ''),
        row.get('Sponsoring 2024', ''),
        nie_mehr_anfragen
    ))
    sponsor_id = cursor.lastrowid
    for mail in emails:
        cursor.execute("INSERT INTO mailadresse (sponsor_id, mail) VALUES (?, ?)", (sponsor_id, mail))

conn.commit()
conn.close()
print("Fertig! Boolesche Werte und Zusatzinfos werden jetzt korrekt verarbeitet.")
